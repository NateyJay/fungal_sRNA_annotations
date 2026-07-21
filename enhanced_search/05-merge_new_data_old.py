
import sys
from pathlib import Path
from os.path import isfile, isdir


from openpyxl import Workbook, load_workbook
import openpyxl.utils.cell
from datetime import datetime

import time
from pprint import pprint


import shutil
import string
import os





row_modes = [
	'allow_new_rows',
	'match_to_input',
	'do_not_change'
]


######################
# Suggested profiles #
######################

### Copying annotation data to the master file

# merge_key = "srr"
# parent_file = '05out-annotated_table.xlsx'
# new_file = "prior_annotations/fungi_libraries.xlsx"
# new_file = "prior_annotations/fabian_table.xlsx"
# row_mode = row_modes[2]

# permitted_columns = ['srr','To Include']


### Updating to a newly processed parent table

merge_key = "srr"
# parent_file = '05out-annotated_table.xlsx'
parent_file = "/Volumes/YASMA/Information_tables/Annotated_libraries.xlsx"
new_file = "03out-output_table.keywords.txt"
row_mode = row_modes[1]
permitted_columns = []














print("reading in new file")

if new_file.endswith(".txt"):

	with open(new_file, 'r') as f:
		header = f.readline().strip().split("\t")
		srrs = set()


		# print([(h,i) for i,h in enumerate(header) if h not in xl_header])

		table_d = {}
		for h in header:
			table_d[h] = {}

		for line in f:
			line = line.strip().split('\t')
			srr = line[header.index(merge_key)]
			srrs.add(srr)
			for i,l in enumerate(line):

				if permitted_columns:
					if header[i] in permitted_columns:
						table_d[header[i]][srr] = l
				else:
					table_d[header[i]][srr] = l

elif new_file.endswith(".xlsx"):

	workbook = load_workbook(filename=new_file)
	sheet = workbook.worksheets[0]
	srrs = set()


	header = [v.value for v in sheet['1']]

	table_d = {}
	for h in header:
		table_d[h] = {}


	for line in sheet.iter_rows():
		line = [l.value for l in line]
		srr = line[header.index(merge_key)]
		srrs.add(srr)
		for i,l in enumerate(line):

				if permitted_columns:
					if header[i] in permitted_columns:
						table_d[header[i]][srr] = l
				else:
					table_d[header[i]][srr] = l




# pprint(table_d)
print()


workbook = load_workbook(filename=parent_file)
sheet = workbook.worksheets[0]


xl_header = [v.value for v in sheet['1']]
xl_srrs = [v.value for v in sheet[string.ascii_uppercase[xl_header.index(merge_key)]][1:]]




master_header = xl_header + [h for h in header if h not in xl_header]

if row_mode == 'allow_new_rows':
	print(f'row_mode: {row_mode} - new rows will be added to table')
	master_srrs = xl_srrs + [s for s in srrs if s not in xl_srrs]
elif row_mode == 'match_to_input':
	print(f'row_mode: {row_mode} - table will have the same rows as the new input')
	master_srrs = srrs
elif row_mode == 'do_not_change':
	print(f'row_mode: {row_mode} - table will have the same rows as the parent file')
	master_srrs = xl_srrs
else:
	sys.exit('Error: row mode not right!')




cols_to_delete = [i for i,m in enumerate(master_header) if not m]
rows_to_delete = [i for i,m in enumerate(master_srrs) if not m]


print(cols_to_delete, "<- bad columns to be deleted")
print(rows_to_delete, "<- bad rows to be deleted")
print()


for i in cols_to_delete[::-1]:
	sheet.delete_cols(i+1, 1)

for i in rows_to_delete[::-1]:
	sheet.delete_rows(i+1, 1)


print(len(xl_header), 'columns in excel')
print(" ", len([h for h in header if h not in xl_header]), 'columns to add')
print(" ", len([h for h in xl_header if h not in header]), 'columns missing in new table')

print()
print(len(xl_srrs), 'rows in excel')
print(" ", len([s for s in srrs if s not in xl_srrs]), 'rows to add')
print(" ", len([s for s in xl_srrs if s not in srrs]), 'rows missing in new table')
print()

input("Press any key to format the table...")


master_header = [m for m in master_header if m]
master_srrs   = [m for m in master_srrs if m]

# print([v.value for v in sheet["1"]])

# sys.exit()

def offset_plot(v1, v2):
	offset = 20
	v1 = str(v1).strip()
	v2 = str(v2).strip()

	len1 = len(v1)
	len2 = len(v2)

	off1 = offset - len1
	if off1 < 0:
		off1 = 0

	print(" "* off1, end='')
	print(v1[:offset], end='')
	if len1 > offset:
		print("...",end='')
	else:
		print("   ")
	print("  ->  ", end='')
	print(v2[:offset], end='')
	if len2 > offset:
		print("...")
	else:
		print("")





offset = 20
for s in master_srrs:
	for hi,h in enumerate(master_header):

		try:
			si = xl_srrs.index(s)
		except ValueError:
			si = len(xl_srrs)
			xl_srrs.append(s)

		if h not in xl_header:
			key = openpyxl.utils.cell.get_column_letter(hi+1) + '1'
			sheet[key].value = h

		key = openpyxl.utils.cell.get_column_letter(hi+1) + str(si+2)
		val = sheet[key].value

		print

		try:
			new_val = table_d[h][s]
		except:
			new_val = None

		if str(new_val) != str(val) and new_val and new_val != "-":

			print()
			print('header:', h)
			print("srr:   ", s)
			print()
			print(key, "<- excel key")
			print(f"{h},{s}", "<- new key")
			print()

			offset_plot(val, new_val)

			sheet[key].value = new_val

			# input()



print()
print("backing up parent table...")

now = datetime.now()
dt_string = now.strftime("%Y.%m.%d")

base = Path(parent_file).name.replace(".xlsx", '')
copy_file = f"05-backup_tables/" + now.strftime("%Y.%m.%d") + "/" + base + "_" + str(time.time()) + ".xlsx"

Path("05-backup_tables/" + now.strftime("%Y.%m.%d")).mkdir(parents=True, exist_ok=True)

print(f"backing up: \n  {parent_file} \n    to \n  {copy_file}")
print()

shutil.copyfile(parent_file, copy_file)

print()
print("overwriting excel file...")

workbook.save("05-temp.xlsx")
os.rename("05-temp.xlsx", "05out-annotated_table.xlsx")

## 












