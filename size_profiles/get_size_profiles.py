import sys
from pathlib import Path

from openpyxl import load_workbook



wb = load_workbook(filename = '/Volumes/YASMA/master_table.xlsx', data_only=True)

ws = wb.active


condition_d = {}

header = ws['2']
header = [v.value for v in header]
for i,j in enumerate(header):
	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	rg         = row[header.index('Replicate group')]

	if rg:
		print(bioproject, srr, rg, sep='\t')


		try:
			condition_d[bioproject].append(f"{srr}:{rg}")
		except KeyError:
			condition_d[bioproject] = [f"{srr}:{rg}"]



command = "yasma3 "









