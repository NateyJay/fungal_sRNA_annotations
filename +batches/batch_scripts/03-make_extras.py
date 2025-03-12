import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

from random import shuffle

run_all_anyways = True

script_dir = Path("03out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("03out-batch_outputs").mkdir(parents=True, exist_ok=True)



wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	rg         = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if rg:
		# print(bioproject, srr, rg, sep='\t')


		try:
			bpj_d[project].append((srr,rg))
		except KeyError:
			bpj_d[project] = [(srr,rg)]

pprint(bpj_d)

def check_done(project, cond):
	log_file = Path("../annotations", project, f"tradeoff_{cond}", "log.txt")
	if not log_file.is_file():
		return False

	with open(log_file, 'r') as f:
		for line in f:
			if "Run completed:" in line:
				return True


config_file = Path("03-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

i=0
projects = list(bpj_d.keys())
projects.sort()

for project in projects:
	libraries = bpj_d[project]
	abbv = project.split(".")[0]


	rgs = [rg for srr,rg in libraries if "." not in rg]
	rgs = list(set(rgs))
	rgs.sort()

	for cond in rgs:

		script_file = Path(script_dir, f"{project}.{cond}.sh")
		
		with open(script_file, 'w') as outf:

			print(f"""#!/bin/bash

mkdir ../annotations/{project}
cd ../annotations/{project}

echo {project}.{cond}


echo "\nsize_profile..."
yasma.py size-profile -o . -a ./align/alignment.bam

# echo "\nhairpin..."
# yasma.py hairpin -o . -n {cond} -a tradeoff_{cond}


echo "\n"



""", file=outf)

		i += 1

		with open(config_file, 'a') as outf:
			print(i, cond, str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)
			

		if check_done(project, cond) and not run_all_anyways:
			continue
		
		print(i, project, cond, sep='\t')
		to_run.append(i)



	print()

shuffle(to_run)

exec_file = Path("03-command.sh")
with open(exec_file, 'w') as outf:
	outf.write(f'''#!/bin/bash
#SBATCH --cpus-per-task=1
#SBATCH --ntasks-per-node 5
#SBATCH --output=03out-batch_outputs/%a.out.txt 
#SBATCH --error=03out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,to_run))}%50

module load python3/3.8.5
module load bowtie
module load viennarna/2.4.15

config=03-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $4}}' $config)


sh $file



''')











