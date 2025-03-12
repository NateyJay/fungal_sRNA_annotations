import sys
from pathlib import Path
import os

from openpyxl import load_workbook

from subprocess import Popen, PIPE


projects = [
"Agbis.PRJNA770841",
"Alalt.PRJNA475703",
"Albra.PRJNA556655",
"Asfla.PRJNA272148",
"Asfla.PRJNA438019",
"Asfla.PRJNA816993",
"Asfum.PRJNA261826",
"Asfum.PRJNA261827",
"Asfum.PRJNA317629",
"Asfum.PRJNA926984",
"Asrab.PRJNA479940",
"Atrol.PRJNA486707",
"Aupul.PRJNA892012",
"Bebas.PRJNA517599",
"Bebas.PRJNA647702",
"Blgra.PRJNA142095",
"Blgra.PRJNA476756",
"Blgra.PRJNA479878",
"Blhor.PRJNA809109",
"Bocin.PRJNA1092616",
"Bocin.PRJNA193535",
"Bocin.PRJNA193536",
"Bocin.PRJNA193537",
"Bocin.PRJNA253747",
"Bocin.PRJNA282704",
"Bocin.PRJNA282705",
"Bocin.PRJNA325479",
"Bocin.PRJNA342517",
"Bocin.PRJNA431815",
"Bocin.PRJNA496584",
"Bocin.PRJNA730711",
"Bocin.PRJNA752615",
"Bocin.PRJNA978613",
"Bodot.PRJNA511629",
"Boell.PRJNA383018",
"Caalb.PRJNA317629",
"Caalb.PRJNA715092",
"Caalb.PRJNA773057",
"Capar.PRJNA715092",
"Clros.PRJEB43636",
"Clros.PRJEB51338",
"Cocin.PRJNA477255",
"Cocin.PRJNA560364",
"Coglo.PRJNA490143",
"Cohig.PRJNA264848",
"Comil.PRJNA496418",
"Cosub.PRJNA667277",
"Crgat.PRJNA368962",
"Crneo.PRJNA185599",
"Crneo.PRJNA526042",
"Crneo.PRJNA629419",
"Culun.PRJNA769949",
"Diseg.PRJNA534364",
"Epnig.PRJNA811108",
"Focan.PRJNA705837",
"Fubra.PRJNA510758",
"Fugra.PRJNA248275",
"Fugra.PRJNA253151",
"Fugra.PRJNA253153",
"Fugra.PRJNA304218",
"Fugra.PRJNA347833",
"Fugra.PRJNA431527",
"Fugra.PRJNA683746",
"Fugra.PRJNA749737",
"Fugra.PRJNA900007",
"Fuoxy.PRJNA232807",
"Fuoxy.PRJNA329032",
"Fuoxy.PRJNA407898",
"Fuoxy.PRJNA472169",
"Fuoxy.PRJNA562097",
"Fuoxy.PRJNA723916",
"Fuoxy.PRJNA760453",
"Gasin.PRJNA42807",
"Gimar.PRJEB35457",
"Glint.PRJNA142331",
"Glint.PRJNA437917",
"Hicap.PRJNA514312",
"Labic.PRJNA481323",
"Lathe.PRJNA558429",
"Mabru.PRJNA731035",
"Maory.PRJNA185495",
"Maory.PRJNA326250",
"Maory.PRJNA504419",
"Maory.PRJNA751253",
"Maory.PRJNA856435",
"Masym.PRJNA342612",
"Meacr.PRJNA688941",
"M.ory.PRJNA310070",
"Mucir.PRJNA200295",
"Mucir.PRJNA453739",
"Mulus.PRJNA243024",
"Mulus.PRJNA903107",
"Nacas.PRJNA354404",
"Nacas.PRJNA530565",
"Nacas.PRJNA530697",
"Nacas.PRJNA600940",
"Necra.PRJNA125805",
"Necra.PRJNA167682",
"Necra.PRJNA190099",
"Necra.PRJNA196947",
"Necra.PRJNA207075",
"Necra.PRJNA350329",
"Nobom.PRJNA760284",
"Nobom.PRJNA953616",
"Nocer.PRJNA408312",
"Nocer.PRJNA487111",
"Nocer.PRJNA562787",
"Opsin.PRJNA673414",
"Pabra.PRJNA480504",
"Pabra.PRJNA931606",
"Peita.PRJNA576793",
"Petra.PRJNA742222",
"Petra.PRJNA756805",
"Plcor.PRJNA944818",
"Plery.PRJNA450159",
"Plost.PRJNA448380",
"Pltuo.PRJNA450159",
"Pugra.PRJNA960906",
"Pustr.PRJNA289147",
"Pustr.PRJNA355964",
"Putri.PRJNA266709",
"Pyory.PRJNA322180",
"Rasol.PRJNA213313",
"Rhint.PRJNA740297",
"Rhirr.PRJEB29180",
"Rhirr.PRJNA213313",
"Rhirr.PRJNA429556",
"Rhirr.PRJNA481323",
"Rhirr.PRJNA722321",
"Rhsol.PRJNA282111",
"Rhsol.PRJNA596921",
"Rhsol.PRJNA900007",
"Rhsp..PRJNA631292",
"Sacer.PRJNA154125",
"Sacer.PRJNA154129",
"Sacer.PRJNA499084",
"Savan.PRJNA798153",
"Scjap.PRJNA770349",
"Scpom.PRJNA120293",
"Scpom.PRJNA122193",
"Scpom.PRJNA125397",
"Scpom.PRJNA144481",
"Scpom.PRJNA154563",
"Scpom.PRJNA168300",
"Scpom.PRJNA177799",
"Scpom.PRJNA183638",
"Scpom.PRJNA229167",
"Scpom.PRJNA235985",
"Scpom.PRJNA254525",
"Scpom.PRJNA259168",
"Scpom.PRJNA259172",
"Scpom.PRJNA273446",
"Scpom.PRJNA278408",
"Scpom.PRJNA301367",
"Scpom.PRJNA321172",
"Scpom.PRJNA322452",
"Scpom.PRJNA322455",
"Scpom.PRJNA341984",
"Scpom.PRJNA350403",
"Scpom.PRJNA350506",
"Scpom.PRJNA361034",
"Scpom.PRJNA369032",
"Scpom.PRJNA378525",
"Scpom.PRJNA382810",
"Scpom.PRJNA383112",
"Scpom.PRJNA438370",
"Scpom.PRJNA557604",
"Scpom.PRJNA575857",
"Scpom.PRJNA727318",
"Scscl.PRJNA140539",
"Scscl.PRJNA315516",
"Scscl.PRJNA348385",
"Scscl.PRJNA361523",
"Scscl.PRJNA379694",
"Scscl.PRJNA477286",
"Scscl.PRJNA607657",
"Scscl.PRJNA659617",
"Scscl.PRJNA678586",
"Scscl.PRJNA746614",
"Spsci.PRJNA322114",
"Tacam.PRJNA268267",
"Tamar.PRJNA207279",
"Tamar.PRJNA647397",
"Trasp.PRJNA638238",
"Tratr.PRJNA508370",
"Trree.PRJNA201504",
"Trrub.PRJNA483837",
"Trrub.PRJNA627692",
"Vacer.PRJNA399493",
"Vamal.PRJNA413773",
"Vamal.PRJNA542139",
"Vapol.PRJNA140091",
"Vealb.PRJNA213313",
"Vedah.PRJNA198742",
"Vedah.PRJNA787244",
"Vedah.PRJNA794992",
"Vedah.PRJNA819185",
"Vedal.PRJNA592621",
"Venon.PRJNA624041",
"Venon.PRJNA665133",
"Vovol.PRJNA594834",
"Yalip.PRJNA964764",
"Zytri.PRJNA271281",
"Zytri.PRJNA480952"]





out_dir = Path("01out-depth_profiles")
out_dir.mkdir(parents=True, exist_ok=True)

for p in projects:

	call = f"scp njohnson@darwin:/home2/njohnson/fungi_annotations/fungi_annotations_august/annotations/{p}/nativealign/alignment.depth.txt {str(out_dir)}/{p}.depth.txt"

	print()
	print(call)

	p = Popen(call, shell=True)
	p.wait()












sys.exit()

wb = load_workbook(filename = '/Volumes/YASMA/master_table.xlsx', data_only=True)

ws = wb.active


condition_d = {}

header = ws['2']
header = [v.value for v in header]
for i,j in enumerate(header):
	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	rg         = row[header.index('Replicate group')]

	if rg:
		print(bioproject, srr, rg, sep='\t')


		try:
			condition_d[bioproject].append(f"{srr}:{rg}")
		except KeyError:
			condition_d[bioproject] = [f"{srr}:{rg}"]



command = "yasma3 "











