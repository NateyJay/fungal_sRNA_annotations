import sys
from pprint import pprint
import json
from subprocess import Popen, PIPE
from pathlib import Path
from openpyxl import load_workbook


wb = load_workbook("Species_information.xlsx")

ws = wb.active


header = [s.value for s in ws['1']]
# for i,j in enumerate(header):
# 	print(i,j)

out_file = "01out-taxonomy_info.txt"
with open(out_file, 'w') as outf:
	print("abbv","species",'genus_tax_id','taxid','phylum','class','order','family','genus', sep='\t', file=outf)


species_i = header.index("Species")
abbv_i    = header.index("Code")


problem_genera = {
	'Cryptococcus': 5206,
	"Rhizophagus" : 1129544,
}


def get_species_taxid(species):
	if len(species.split()) > 2:
		species = " ".join(species.split()[:2])


	call = ['datasets', 'summary', 'taxonomy', 'taxon', species]

	p = Popen(call, stdout=PIPE, stderr=PIPE, encoding='utf-8')

	out, err = p.communicate()

	try:
		data = json.loads(out)
	except json.decoder.JSONDecodeError:
		print(species, "<- species error")
	# sys.exit()
	taxid  = data['reports'][0]['taxonomy']['tax_id']
	return(taxid)


for row in ws.iter_rows(values_only=True, min_row=2):
	species = row[species_i]
	abbv    = row[abbv_i]
	genus   = species.split()[0]

	# print(species)



	try:
		call = ['datasets', 'summary', 'taxonomy', 'taxon', str(problem_genera[genus])]

	except KeyError:
		call = ['datasets', 'summary', 'taxonomy', 'taxon', genus]

	p = Popen(call, stdout=PIPE, stderr=PIPE, encoding='utf-8')

	out, err = p.communicate()

	try:
		data = json.loads(out)
	except json.decoder.JSONDecodeError:
		print(species, "<- error")
		continue
		# sys.exit()

	species_taxid = get_species_taxid(species)


	taxid  = data['reports'][0]['taxonomy']['tax_id']
	phylum = data['reports'][0]['taxonomy']['classification']['phylum']['name']

	try:
		classs = data['reports'][0]['taxonomy']['classification']['class']['name']
	except KeyError:
		classs = "NONE"

	try:
		orderr = data['reports'][0]['taxonomy']['classification']['order']['name']
	except KeyError:
		orderr = "NONE"
	try:
		family = data['reports'][0]['taxonomy']['classification']['family']['name']
	except KeyError:
		family = "NONE"

	try:
		genus  = data['reports'][0]['taxonomy']['classification']['genus']['name']
	except KeyError:
		genus = "NONE"




	# print(phylum)

	with open(out_file, 'a') as outf:
		print(abbv, species, taxid, species_taxid, phylum, classs, orderr, family, genus, sep='\t', file=outf)
		print(abbv, taxid, phylum, species, sep='\t')

	# sys.exit()

# with open('/path/to/file.jsonl') as f:
#     data = [json.loads(line) for line in f]
