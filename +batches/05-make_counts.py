import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

from random import shuffle

run_all_anyways = True

script_dir = Path("05out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("05out-batch_outputs").mkdir(parents=True, exist_ok=True)



wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	rg         = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if rg:
		# print(bioproject, srr, rg, sep='\t')

		if project in bpj_d:
			if rg not in bpj_d[project]:
				bpj_d[project].append(rg)
		else:
			bpj_d[project] = [rg]

pprint(bpj_d)

def check_done(project):
	return(False)
	log_file = Path("../annotations", project, f"counts", f"tradeoff_deepcounts.txt")
	if not log_file.is_file():
		return False

	with open(log_file, 'r') as f:
		for line in f:
			if "Run completed:" in line:
				return True


config_file = Path("05-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

i=0
projects = list(bpj_d.keys())
projects.sort()

# print(len(bpj_d))

for project in projects:
	libraries = bpj_d[project]
	abbv = project.split(".")[0]

	script_file = Path(script_dir, f"{project}.sh")
	
	with open(script_file, 'w') as outf:

		print(f"""#!/bin/bash

source /home/opt/anaconda3/etc/profile.d/conda.sh
conda activate nate_env


mkdir ../annotations/{project}
cd ../annotations/{project}

echo {project}

echo "counting..."
yasma.py count -o . -an */loci.gff3 -an ../../TE/{abbv}.filteredRepeats.gff -an ../../rfam/{abbv}.rfam.gff3 -an ../../metaloci/{abbv}.meta.gff3

""", file=outf)

		i += 1

		with open(config_file, 'a') as outf:
			print(i, "NA", str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)
			

		if check_done(project) and not run_all_anyways:
			continue
		
		print(i, project,  sep='\t')
		to_run.append(i)




shuffle(to_run)
max_jobs = 500
i=0
while i <= len(to_run):
	jobs = to_run[i: i+500]



	exec_file = Path(f"05-command_{i}.sh")
	with open(exec_file, 'w') as outf:
		outf.write(f'''#!/bin/bash
#SBATCH --cpus-per-task=1
#SBATCH --ntasks-per-node 5
#SBATCH --output=05out-batch_outputs/%a.out.txt 
#SBATCH --error=05out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,jobs))}%50


config=05-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $4}}' $config)


sh $file



''')

	i += max_jobs












