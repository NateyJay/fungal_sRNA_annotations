
import sys
from pathlib import Path
from os.path import isfile, isdir


from openpyxl import Workbook, load_workbook
import openpyxl.utils.cell
from datetime import datetime

import time
from pprint import pprint


import shutil
import string
import os





key = 'pmid'



def check_txt(file):
	print(" ", file)
	bpjs = set()
	with open(file, 'r') as f:
		header = f.readline()

		header = header.strip().split("\t")

		for line in f:
			bpjs.add(line.strip().split("\t")[header.index(key)])

	return(bpjs)


def check_excel(file):

	print(" ",file)
	workbook = load_workbook(filename=file)
	sheet = workbook.worksheets[0]


	header = [v.value for v in sheet['1']]

	# print(header)

	column_letter = openpyxl.utils.cell.get_column_letter(header.index(key)+1)

	bpjs = sheet[column_letter]
	bpjs = [b.value for b in bpjs]
	bpjs = set(list(bpjs[1:]))

	return(bpjs)




bpj_sets = dict()

# bpj_sets['fabian_table'] = check_excel("04-fabian_table.xlsx")
# bpj_sets['annotated_table'] = check_excel("04out-annotated_table.xlsx")
# bpj_sets['original_set'] = check_excel("prior_annotations/fungi_libraries.xlsx")
bpj_sets['bbernal_table'] = check_excel("prior_annotations/bbernal_srna_fungi_data.xlsx")
bpj_sets['fresh_table'] = check_txt("03out-output_table.keywords.txt")

print()

master_set = set()
for bpjs in bpj_sets.values():
	master_set.update(bpjs)

with open("04out-checked.txt", 'w') as outf:
	header = [key] + [n for n in bpj_sets.keys()]
	print("\t".join(header), file=outf)
	print("\t".join(header))
	for bpj in master_set:
		line = [bpj]
		for bpjs in bpj_sets.values():
			line.append(str(int(bpj in bpjs)))
		print("\t".join(map(str,line)), file=outf)
		print("\t".join(map(str,line)))


sys.exit()













