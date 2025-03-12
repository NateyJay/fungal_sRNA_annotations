
import sys
import pysam
from pprint import pprint
from collections import Counter
from openpyxl import load_workbook
from pathlib import Path



## scp 01-get_raw_alignments.py njohnson@darwin:/home2/njohnson/fungi_annotations/mRNA_alignment/
## scp njohnson@darwin:/home2/njohnson/fungi_annotations/mRNA_alignment/01out-merged.txt ./



def check_coarse_overlap(gff, bam):



	ann_d = {}

	if gff.is_file():

		with open(gff, 'r') as f:

			for line in f:

				if line.startswith("#"):
					continue

				line = line.strip().split("\t")

				if line[2] != 'mRNA':
					continue

				strand = line[6]

				for r in range(int(line[3]), int(line[4])):
					try:
						ann_d[line[0]][r] = strand
					except KeyError:
						ann_d[line[0]] = {r : strand}

				# ann_d[line[0]][r] = strand
				

	bamf = pysam.AlignmentFile(bam, "rb")	

	if not bamf.has_index():
		print(f'   index not found for {bam}. Indexing with samtools.')

		pysam.index(str(bam))
		bamf.close()
		bamf = pysam.AlignmentFile(bam,'rb')



	c = Counter()






	for read in bamf.fetch(until_eof=True):

		# print(read)
		# sys.exit()

		read_strand = "+" if read.is_forward else "-"

		# print([read.reference_name, read.reference_start])

		# print(gff_d[read.reference_name])
		# sys.exit()
		try:
			start_overlap = ann_d[read.reference_name][read.reference_start]
		except KeyError:
			start_overlap = False

		try:
			end_overlap = ann_d[read.reference_name][read.reference_end]
		except KeyError:
			end_overlap = False

		# print(read.reference_name)
		# print(read.reference_start)
		# print(read.reference_end)
		# print(start_overlap)
		# print(end_overlap)

		# input()



		if not read.is_mapped:
			key = ('unaligned', read.query_length)

		elif start_overlap == False and end_overlap == False:
			key = ('aligned', read.query_length)

		elif start_overlap == read_strand or end_overlap == read_strand:
			key = ('gene_aligned_stranded', read.query_length)

		else:
			key = ('gene_aligned', read.query_length)

		c[key] += 1




	bamf.close()

	return(c)


wb = load_workbook(filename = '../batch_scripts/master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

print('reading master file...', flush=True)
for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	cond       = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if cond:

		try:
			bpj_d[project].append((srr,cond))
		except KeyError:
			bpj_d[project] = [(srr,cond)]


projects = list(bpj_d.keys())
projects.sort()


print("getting gff directories...", flush=True)
gffs = list(Path("../Genomes/").glob('*.gff*'))
gff_d = {}
for gff in gffs:
	abbv = gff.name.split(".")[0]
	gff_d[abbv] = gff


print('checking for overlaps...', flush=True)

# for project in ['Bocin.PRJNA325479']:
for i,project in enumerate(projects):


	abbv = project[:5]

	print(f"  {i+1}/{len(projects)}\t{project}", flush=True)

	out_file = Path(f"01out-{project}.txt")

	if out_file.is_file() and out_file.stat().st_size > 1000:
		print("    ^^^ done already, skip...")
		continue



	try:
		gff = gff_d[abbv]
	except KeyError:
		print("    ^^^ no gff found (continuing with only alignments)", flush = True)
		# continue

	if not gff.is_file():
		print("    ^^^ no gff found (continuing with only alignments)", flush = True)
		# continue

	try:
		f = open(gff,'r')
		f.close()
	except PermissionError:
		print("    ^^^ bad gff permissions", flush = True)
		continue


	bam = Path("../annotations", project, 'align/alignment.bam')
	if not bam.is_file():
		print("    ^^^ no bam found", flush = True)
		continue


	c = check_coarse_overlap(gff, bam)


	with open(out_file, 'w') as outf:
		# print("project", "cat", 'size','abd', sep='\t', file=outf)
		print('project','size', 'unaligned','aligned','gene_aligned','gene_aligned_stranded', sep='\t', file=outf, flush=True)
		for size in range(15, 51):
			print(project, size, sep='\t', end='', file=outf)
			for cat in ['unaligned','aligned','gene_aligned','gene_aligned_stranded']:
				print('\t' + str(c[(cat, size)]), end='', file=outf)
			print('', end='\n', flush=True, file=outf)



with open('01out-merged.txt', 'w') as outf:
	print('project','size', 'unaligned','aligned','gene_aligned','gene_aligned_stranded', sep='\t', file=outf, flush=True)

	for project in projects:

		out_file = Path(f"01out-{project}.txt")

		if not out_file.is_file() or out_file.stat().st_size < 1000:
			continue

		with open(out_file, 'r') as f:
			f.readline()
			for line in f:
				print(line.strip(), file=outf)










