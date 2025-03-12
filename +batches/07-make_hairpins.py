import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

from random import shuffle

run_all_anyways = True

script_dir = Path("07out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("07out-batch_outputs").mkdir(parents=True, exist_ok=True)


genomes = {}
genome_dir = Path("../Genomes")

for g in genome_dir.iterdir():
	abbv = g.name.split(".")[0]

	if g.suffix == '.fa':
		genomes[abbv] = Path("../../Genomes", g.name)


wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	cond       = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if cond:
		# print(bioproject, srr, rg, sep='\t')


		try:
			bpj_d[project].add(cond)
		except KeyError:
			bpj_d[project] = set([cond])

pprint(bpj_d)

def check_done(project):
	return(False)
	log_file = Path("../annotations", project, f"hairpin", "log.txt")
	if not log_file.is_file():
		return False

	with open(log_file, 'r') as f:
		for line in f:
			if "Run completed:" in line:
				return True


config_file = Path("07-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

i=0
projects = list(bpj_d.keys())
projects.sort()

print(len(projects))

for project in projects:
	abbv = project.split(".")[0]

	try:
		genome = genomes[abbv]
	except KeyError:
		# print(project, libraries)
		continue


	conditions = bpj_d[project]

	for cond in conditions:

		if "." in cond:
			continue

		script_file = Path(script_dir, f"{project}.{cond}.sh")
		
		with open(script_file, 'w') as outf:

			print(f"""#!/bin/bash


	source /home/opt/anaconda3/etc/profile.d/conda.sh
	conda activate nate_env

	mkdir ../annotations/{project}
	cd ../annotations/{project}

	echo {project}


	echo "hairpin..."
	yasma.py hairpin -o . -a align/alignment.bam --annotation_folder tradeoff_{cond} -g {genome} --name {cond} --cores 56




	""", file=outf)

		i += 1

		with open(config_file, 'a') as outf:
			print(i, cond, str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)
			

		if check_done(project) and not run_all_anyways:
			continue
		
		print(i, project, cond, sep='\t')
		to_run.append(i)




# shuffle(to_run)

exec_file = Path("07-command.sh")
with open(exec_file, 'w') as outf:
	outf.write(f'''#!/bin/bash
#SBATCH --cpus-per-task=28
#SBATCH --ntasks-per-node 1
#SBATCH --output=07out-batch_outputs/%a.out.txt 
#SBATCH --error=07out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,to_run))}%6

module load bowtie

config=07-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $4}}' $config)


sh $file



''')











