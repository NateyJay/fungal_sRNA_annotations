import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

from random import shuffle

run_all_anyways = True

script_dir = Path("02out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("02out-batch_outputs").mkdir(parents=True, exist_ok=True)



wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	cond       = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if cond:

		try:
			bpj_d[project].append((srr,cond))
		except KeyError:
			bpj_d[project] = [(srr,cond)]

pprint(bpj_d)

config_file = Path("02-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

i=0
projects = list(bpj_d.keys())
projects.sort()

for project in projects:
	libraries = bpj_d[project]
	abbv = project.split(".")[0]

	condition_str = [f"{srr}:{cond}" for srr, cond in libraries if "." not in cond]
	annotation_conditions = [cond for srr, cond in libraries if "." not in cond]

	def f7(seq):
		seen = set()
		seen_add = seen.add
		return [x for x in seq if not (x in seen or seen_add(x))]

	annotation_conditions = f7(annotation_conditions)

	condition_str = " ".join(condition_str)
	annotation_conditions_str = " ".join(annotation_conditions)

	# print(annotation_conditions)
	# rgs = [rg for srr,rg in libraries if "." not in rg]
	# rgs = list(set(rgs))
	# rgs.sort()

	# for cond in rgs:

	script_file = Path(script_dir, f"{project}.sh")
	
	with open(script_file, 'w') as outf:

		print(f"""#!/bin/bash

mkdir ../annotations/{project}
cd ../annotations/{project}

echo {project}


echo "\ncopying alignment to scratch..."
cp ./align/alignment.bam /scratch/njohnson/{project}.bam

echo "\nannotating..."
yasma.py tradeoff -o . -ac {annotation_conditions_str} -c {condition_str} -a /scratch/njohnson/{project}.bam


echo "\ndeleting scratch alignment..."
rm /scratch/njohnson/{project}.bam
rm /scratch/njohnson/{project}.bam.bai
rm /scratch/njohnson/{project}.depth.txt


""", file=outf)


	def check_done(project, conditions):

		# print(conditions)

		# print("_".join(conditions))
		# sys.exit()
		# wrong_folder = Path("../annotations", project, f"tradeoff_" + "_".join(conditions))
		# right_folder = Path("../annotations", project, f"tradeoff")
		# # print(wrong_folder)
		# if wrong_folder.is_dir():
		# 	print(wrong_folder)
		# 	wrong_folder.rename(right_folder)
		# 	print(right_folder)
		# 	# print(wrong_folder)


			# sys.exit()


		log_file = Path("../annotations", project, "tradeoff", "log.txt")
		if not log_file.is_file():

			# print(right_folder)
			# print(wrong_folder. is_dir(), "<- wrong_folder", wrong_folder)
			# print(right_folder.is_dir(), "<- right_folder", right_folder)
			# print(log_file)
			# print("lf not found")
			# print()
			return False



		with open(log_file, 'r') as f:
			for line in f:
				if "Run completed:" in line:
					return True




	i += 1

	with open(config_file, 'a') as outf:
		print(i, None, str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)
		

	if check_done(project, annotation_conditions) and not run_all_anyways:
		continue
	
	print(i, project, ",".join(annotation_conditions), sep='\t')
	to_run.append(i)




shuffle(to_run)

exec_file = Path("02-command.sh")
with open(exec_file, 'w') as outf:
	outf.write(f'''#!/bin/bash
#SBATCH --cpus-per-task=1
#SBATCH --ntasks-per-node 5
#SBATCH --output=02out-batch_outputs/%a.out.txt 
#SBATCH --error=02out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,to_run))}%50

module load python3/3.8.5
module load bowtie

config=02-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $4}}' $config)


sh $file



''')



print(len(to_run), 'out of', len(projects), 'to run')







