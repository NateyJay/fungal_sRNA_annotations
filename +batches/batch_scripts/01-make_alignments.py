import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

run_all_anyways = True


genomes = {}
genome_dir = Path("../Genomes")

for g in genome_dir.iterdir():
	abbv = g.name.split(".")[0]

	if g.suffix == '.fa':
		genomes[abbv] = Path("../../Genomes", g.name)


script_dir = Path("01out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("01out-batch_outputs").mkdir(parents=True, exist_ok=True)



wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	rg         = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if rg:
		# print(bioproject, srr, rg, sep='\t')


		try:
			bpj_d[project].append((srr,rg))
		except KeyError:
			bpj_d[project] = [(srr,rg)]

pprint(bpj_d)

def check_done(project):

	if run_all_anyways:
		print("RAA")
		return False
		
	log_file = Path("../annotations", project, "align", "log.txt")
	if not log_file.is_file():
		return False

	with open(log_file, 'r') as f:
		for line in f:
			if "alignment complete!" in line:
				return True


config_file = Path("01-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

for i, project in enumerate(bpj_d.keys()):
	libraries = bpj_d[project]
	abbv = project.split(".")[0]
	try:
		genome = genomes[abbv]
	except KeyError:
		# print(project, libraries)
		continue


	srr_string = [srr for srr,rg in libraries]
	srr_string = " ".join(srr_string)

	cond_string = [f"{srr}:{rg}" for srr,rg in libraries]
	cond_string = " ".join(cond_string)

	script_file = Path(script_dir, f"{project}.sh")
	
	with open(script_file, 'w') as outf:

		print(f"""#!/bin/bash

mkdir ../annotations/{project}
cd ../annotations/{project}

echo {project}


echo "\nmaking inputs..."
yasma.py inputs \\
-o . \\
--srrs {srr_string} \\
--conditions {cond_string} \\
--genome_file {genome}

echo "\ndownloading..."
yasma.py download -o .

echo "\nfinding adapters..."
yasma.py adapter -o .

echo "\ntrimming..."
yasma.py trim -o . --cores 28

echo "\naligning..."
yasma.py align -o . --cores 28

""", file=outf)


	if check_done(project):
		continue

	with open(config_file, 'a') as outf:
		print(i, str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)

		to_run.append(i)

		print(i, project, sep='\t')


exec_file = Path("01-command.sh")
with open(exec_file, 'w') as outf:
	outf.write(f'''#!/bin/bash
#SBATCH --nodes=1
#SBATCH --cpus-per-task=28
#SBATCH --mem=62GB
#SBATCH --output=01out-batch_outputs/%a.out.txt 
#SBATCH --error=01out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,to_run))}%10

module load python3/3.8.5
module load bowtie

config=01-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $2}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)


sh $file



''')











