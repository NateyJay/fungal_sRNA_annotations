import sys
from pprint import pprint

from subprocess import Popen, PIPE
from pathlib import Path

from openpyxl import load_workbook

from random import shuffle

run_all_anyways = False

script_dir = Path("02out-scripts")
script_dir.mkdir(parents=True, exist_ok=True)

Path("02out-batch_outputs").mkdir(parents=True, exist_ok=True)



wb = load_workbook(filename = 'master_table.xlsx', data_only=True)

ws = wb.active

bpj_d = {}

header = ws['2']
header = [v.value for v in header]
# for i,j in enumerate(header):
# 	print(i,j)

for i,row in enumerate(ws.iter_rows(min_row = 3, values_only=True)):

	bioproject = row[header.index('bioproject')]
	srr        = row[header.index('srr')]
	cond       = row[header.index('Replicate group')]
	abbv       = row[header.index('fungi_abbv')]

	project = f"{abbv}.{bioproject}"


	if cond:

		try:
			bpj_d[project].append((srr,cond))
		except KeyError:
			bpj_d[project] = [(srr,cond)]

pprint(bpj_d)

config_file = Path("02-config.txt")
with open(config_file, 'w') as outf:
	outf.write('ArrayTaskID\tProject\tAnnDir\n')



print("")
print("To run:")
to_run = []

i=0
projects = list(bpj_d.keys())
projects.sort()

done = 0
not_ready = 0
total_jobs = 0 

for project in projects:
	libraries = bpj_d[project]
	abbv = project.split(".")[0]

	condition_str = [f"{srr}:{cond}" for srr, cond in libraries if "." not in cond]

	annotation_conditions = [cond for srr, cond in libraries if "." not in cond]

	def f7(seq):
		seen = set()
		seen_add = seen.add
		return [x for x in seq if not (x in seen or seen_add(x))]

	annotation_conditions = f7(annotation_conditions)

	condition_str = " ".join(condition_str)


	for cond in annotation_conditions:
		total_jobs += 1

	# annotation_conditions_str = " ".join(annotation_conditions)

		# print(annotation_conditions)
		# rgs = [rg for srr,rg in libraries if "." not in rg]
		# rgs = list(set(rgs))
		# rgs.sort()

		# for cond in rgs:

		script_file = Path(script_dir, f"{project}.{cond}.sh")
		
		with open(script_file, 'w') as outf:

			print(f"""#!/bin/bash

mkdir ../annotations/{project}
cd ../annotations/{project}	


source /home/opt/anaconda3/etc/profile.d/conda.sh
conda activate nate_env

echo {project}.{cond}


echo "\nannotating..."
yasma.py tradeoff -o . -ac {cond} -c {condition_str} -a align/alignment.bam -n {cond}




""", file=outf)



		def check_done(project):

			log_file = Path("../annotations", project, f"tradeoff_{cond}", "log.txt")
			if not log_file.is_file():
				return False

			with open(log_file, 'r') as f:
				for line in f:
					if "Run completed:" in line:
						return True

		def check_ready(project):
			alignment_file = Path("../annotations", project, 'align', 'alignment.bam')
			log_file = Path("../annotations", project, "align", "log.txt")

			if not log_file.is_file():
				return False

			if not alignment_file.is_file():
				return False

			# print(alignment_file.stat().st_size)
			if alignment_file.stat().st_size > 1000000:
				return True
			else:
				return False

			with open(log_file, 'r') as f:
				for line in f:
					if "alignment complete!" in line:
						return True

			return False



			with open(log_file, 'r') as f:
				for line in f:
					if "Run completed:" in line:
						return True




		i += 1

		with open(config_file, 'a') as outf:
			print(i, cond, str(script_file), project, f"../annotations/{project}", sep='\t', file=outf)
			

		# if check_done(project, annotation_conditions) and not run_all_anyways:
		# 	continue
		if run_all_anyways:
			print("run_all_anyways...")
		else:

			if check_done(project):
				done += 1
				print(i, f"done -> {project}.{cond}")
				continue

			if not check_ready(project):
				not_ready += 1
				print(i, f"not_ready -> {project}.{cond}")
				continue
		
		print(i, project, cond, sep='\t')
		to_run.append(i)




shuffle(to_run)

exec_file = Path("02-command.sh")
with open(exec_file, 'w') as outf:
	outf.write(f'''#!/bin/bash
#SBATCH --cpus-per-task=1
#SBATCH --ntasks-per-node 1
#SBATCH --mem=60GB
#SBATCH --output=02out-batch_outputs/%a.out.txt 
#SBATCH --error=02out-batch_outputs/%a.err.txt 
#SBATCH --array {",".join(map(str,to_run))}%4

module load bowtie


config=02-config.txt
file=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $3}}' $config)
project=$(awk -v ArrayTaskID=$SLURM_ARRAY_TASK_ID '$1==ArrayTaskID {{print $4}}' $config)


sh $file



''')


print()
print(f"{done} annotations done")
print(f"{not_ready} alignments not ready")
print(len(to_run), 'out of', total_jobs, 'to run')






